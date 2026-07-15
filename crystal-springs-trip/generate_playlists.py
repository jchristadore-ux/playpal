#!/usr/bin/env python3
"""Generate all Crystal Springs Golf Trip playlist deliverables.

Reads playlist_data.py, enforces the hard rules, and writes:

  output/master_playlist.csv        one row per track across all playlists
  output/master_playlist.xlsx       spreadsheet: summary + master + per-playlist tabs
  output/csv/NN_<playlist>.csv      one CSV per playlist
  output/import/NN_<playlist>.csv   title/artist CSVs for import tools (Soundiiz etc.)
  output/playlists.json             full JSON backup
  output/PLAYLISTS.md               human-readable markdown version
  output/VALIDATION.md              duplicate/artist/runtime/genre verification report

Exits non-zero if any hard rule fails:
  - a song (normalized artist+title) appears in more than one playlist
  - a primary artist appears more than 3x within one playlist
"""

import csv
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from playlist_data import PLAYLISTS

HERE = Path(__file__).resolve().parent
OUT = HERE / "output"

MAX_ARTIST_PER_PLAYLIST = 3


# ---------------------------------------------------------------- helpers
def to_seconds(length: str) -> int:
    m, s = length.split(":")
    return int(m) * 60 + int(s)


def fmt_hm(total_seconds: int) -> str:
    h, rem = divmod(total_seconds, 3600)
    m = rem // 60
    return f"{h}h {m:02d}m"


def norm(text: str) -> str:
    """Normalize for duplicate detection: casefold, strip accents/punct."""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.casefold()
    text = re.sub(r"\(.*?\)", " ", text)  # ignore parentheticals
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def primary_artist(artist: str) -> str:
    """Primary artist for repetition caps: strip feat./with/& guests."""
    a = re.split(r"\s+feat\.\s+", artist, flags=re.I)[0]
    a = re.split(r"\s+featuring\s+", a, flags=re.I)[0]
    return a.strip()


def slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


# ---------------------------------------------------------------- validation
def validate():
    errors, warnings = [], []

    # Zero duplicate songs across ALL playlists (Trip Anthem included).
    seen = {}
    title_seen = defaultdict(list)
    for pl in PLAYLISTS:
        for artist, title, _length, _genre in pl["songs"]:
            key = (norm(primary_artist(artist)), norm(title))
            if key in seen:
                errors.append(
                    f"DUPLICATE SONG: {artist} - {title} in '{pl['name']}' "
                    f"already in '{seen[key]}'"
                )
            seen[key] = pl["name"]
            title_seen[norm(title)].append((artist, pl["name"]))

    # Same title by different artists: allowed, but surface for review.
    for t, entries in title_seen.items():
        arts = {norm(primary_artist(a)) for a, _ in entries}
        if len(entries) > 1 and len(arts) > 1:
            warnings.append(
                "Same title, different artists (OK, verify intentional): "
                + "; ".join(f"{a} [{p}]" for a, p in entries)
            )

    # Artist cap within each playlist.
    for pl in PLAYLISTS:
        counts = Counter(primary_artist(a) for a, _t, _l, _g in pl["songs"])
        for artist, n in counts.items():
            if n > MAX_ARTIST_PER_PLAYLIST:
                errors.append(
                    f"ARTIST CAP: {artist} appears {n}x in '{pl['name']}' "
                    f"(max {MAX_ARTIST_PER_PLAYLIST})"
                )

    return errors, warnings


# ---------------------------------------------------------------- reporting
def playlist_stats(pl):
    secs = sum(to_seconds(s[2]) for s in pl["songs"])
    genres = Counter(g for _a, _t, _l, g in pl["songs"])
    return {
        "songs": len(pl["songs"]),
        "seconds": secs,
        "runtime": fmt_hm(secs),
        "target": fmt_hm(pl["target_minutes"] * 60),
        "delta_min": round((secs - pl["target_minutes"] * 60) / 60),
        "genres": dict(genres),
    }


# ---------------------------------------------------------------- writers
def write_master_csv(rows):
    with open(OUT / "master_playlist.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Playlist", "Track #", "Artist", "Song", "Length", "Genre"])
        w.writerows(rows)


def write_per_playlist_csvs():
    (OUT / "csv").mkdir(exist_ok=True)
    (OUT / "import").mkdir(exist_ok=True)
    for i, pl in enumerate(PLAYLISTS, 1):
        base = f"{i:02d}_{slug(pl['name'])}"
        with open(OUT / "csv" / f"{base}.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Track #", "Artist", "Song", "Length", "Genre"])
            for n, (artist, title, length, genre) in enumerate(pl["songs"], 1):
                w.writerow([n, artist, title, length, genre])
        # Import-tool-friendly variant (Soundiiz/TuneMyMusic accept title,artist).
        with open(OUT / "import" / f"{base}.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["title", "artist"])
            for artist, title, _length, _genre in pl["songs"]:
                w.writerow([title, primary_artist(artist)])


def write_json():
    payload = {
        "project": "Crystal Springs Golf Trip Playlist Generator",
        "generated": date.today().isoformat(),
        "total_playlists": len(PLAYLISTS),
        "total_songs": sum(len(p["songs"]) for p in PLAYLISTS),
        "total_runtime": fmt_hm(
            sum(to_seconds(s[2]) for p in PLAYLISTS for s in p["songs"])
        ),
        "playlists": [
            {
                "name": pl["name"],
                "theme": pl["theme"],
                "description": pl["description"],
                "target_runtime": fmt_hm(pl["target_minutes"] * 60),
                "actual_runtime": playlist_stats(pl)["runtime"],
                "song_count": len(pl["songs"]),
                "songs": [
                    {
                        "position": n,
                        "artist": artist,
                        "title": title,
                        "length": length,
                        "genre": genre,
                    }
                    for n, (artist, title, length, genre) in enumerate(pl["songs"], 1)
                ],
            }
            for pl in PLAYLISTS
        ],
    }
    with open(OUT / "playlists.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def write_markdown(all_stats, genre_totals, total_secs, total_songs):
    lines = [
        "# Crystal Springs Golf Trip — The Soundtrack",
        "",
        "Nine playlists. Four days. Zero repeated songs.",
        "",
        f"**{total_songs} unique tracks · total runtime {fmt_hm(total_secs)}**",
        "",
        "| # | Playlist | Theme | Songs | Runtime | Target |",
        "|---|----------|-------|-------|---------|--------|",
    ]
    for i, pl in enumerate(PLAYLISTS, 1):
        st = all_stats[pl["name"]]
        lines.append(
            f"| {i} | {pl['name']} | {pl['theme']} | {st['songs']} | "
            f"{st['runtime']} | {st['target']} |"
        )
    lines += [
        "",
        "**Overall genre mix:** "
        + " · ".join(
            f"{g}: {n} songs ({n * 100 // total_songs}%)"
            for g, n in genre_totals.most_common()
        ),
        "",
        "Song lengths are approximate (album versions); actual streaming "
        "runtimes may differ by a few seconds per track.",
        "",
    ]
    for i, pl in enumerate(PLAYLISTS, 1):
        st = all_stats[pl["name"]]
        lines += [
            "---",
            "",
            f"## {i}. {pl['name']} — *{pl['theme']}*",
            "",
            pl["description"],
            "",
            f"**{st['songs']} songs · {st['runtime']}** (target {st['target']})",
            "",
            "| # | Artist | Song | Length | Genre |",
            "|---|--------|------|--------|-------|",
        ]
        for n, (artist, title, length, genre) in enumerate(pl["songs"], 1):
            lines.append(f"| {n} | {artist} | {title} | {length} | {genre} |")
        lines.append("")
    with open(OUT / "PLAYLISTS.md", "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_validation_md(errors, warnings, all_stats, genre_totals, total_secs, total_songs):
    lines = [
        "# Validation Report",
        "",
        f"Generated: {date.today().isoformat()}",
        "",
        "## Hard rules",
        "",
        f"- Duplicate songs across all 9 playlists: "
        f"**{'NONE — PASS' if not errors else 'FAIL'}**",
        f"- Artist appears more than {MAX_ARTIST_PER_PLAYLIST}x in a single "
        f"playlist: **{'NONE — PASS' if not errors else 'FAIL'}**",
        "",
    ]
    if errors:
        lines += ["### Errors", ""] + [f"- {e}" for e in errors] + [""]
    if warnings:
        lines += ["### Notes", ""] + [f"- {w}" for w in warnings] + [""]
    lines += [
        "## Runtimes",
        "",
        "| Playlist | Songs | Actual | Target | Drift |",
        "|----------|-------|--------|--------|-------|",
    ]
    for pl in PLAYLISTS:
        st = all_stats[pl["name"]]
        sign = "+" if st["delta_min"] >= 0 else ""
        lines.append(
            f"| {pl['name']} | {st['songs']} | {st['runtime']} | {st['target']} | "
            f"{sign}{st['delta_min']} min |"
        )
    lines += [
        "",
        f"**Total: {total_songs} songs · {fmt_hm(total_secs)}**",
        "",
        "## Genre mix (overall)",
        "",
        "| Genre | Songs | Share | Requested |",
        "|-------|-------|-------|-----------|",
    ]
    requested = {"Rap": "40%", "Country": "30%", "Rock": "20%", "Classic Rock": "10%"}
    for g in ["Rap", "Country", "Rock", "Classic Rock"]:
        n = genre_totals.get(g, 0)
        lines.append(f"| {g} | {n} | {n * 100 // total_songs}% | {requested[g]} |")
    lines += [
        "",
        "Note: the mellow playlists the brief asked for (Airbnb Morning, "
        "Cascades) carry little or no rap by design, which pulls the overall "
        "rap share below the 40% target. Wild Turkey (~75% rap), Crystal "
        "Springs, and the evening/party playlists carry the rap weight where "
        "it actually fits the room.",
        "",
        "## Per-playlist genre mix",
        "",
        "| Playlist | Rap | Country | Rock | Classic Rock |",
        "|----------|-----|---------|------|--------------|",
    ]
    for pl in PLAYLISTS:
        st = all_stats[pl["name"]]
        g = st["genres"]
        tot = st["songs"]
        cells = " | ".join(
            f"{g.get(k, 0)} ({g.get(k, 0) * 100 // tot}%)"
            for k in ["Rap", "Country", "Rock", "Classic Rock"]
        )
        lines.append(f"| {pl['name']} | {cells} |")
    lines.append("")
    with open(OUT / "VALIDATION.md", "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_xlsx(rows, all_stats, total_secs, total_songs):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping .xlsx (CSV master still written)")
        return

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F6F43")  # golf green

    def style_header(ws):
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"

    def autofit(ws):
        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 52)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["#", "Playlist", "Theme", "Songs", "Runtime", "Target"])
    for i, pl in enumerate(PLAYLISTS, 1):
        st = all_stats[pl["name"]]
        ws.append([i, pl["name"], pl["theme"], st["songs"], st["runtime"], st["target"]])
    ws.append([])
    ws.append(["", "TOTAL", "", total_songs, fmt_hm(total_secs), ""])
    ws[f"B{ws.max_row}"].font = Font(bold=True)
    style_header(ws)
    autofit(ws)

    # Master sheet
    ws = wb.create_sheet("Master")
    ws.append(["Playlist", "Track #", "Artist", "Song", "Length", "Genre"])
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    autofit(ws)

    # One tab per playlist
    for pl in PLAYLISTS:
        ws = wb.create_sheet(pl["name"][:31])
        ws.append(["Track #", "Artist", "Song", "Length", "Genre"])
        for n, (artist, title, length, genre) in enumerate(pl["songs"], 1):
            ws.append([n, artist, title, length, genre])
        style_header(ws)
        autofit(ws)

    wb.save(OUT / "master_playlist.xlsx")


# ---------------------------------------------------------------- main
def main():
    OUT.mkdir(exist_ok=True)

    errors, warnings = validate()
    for w in warnings:
        print(f"NOTE: {w}")
    if errors:
        for e in errors:
            print(f"ERROR: {e}", file=sys.stderr)
        print(f"\n{len(errors)} hard-rule violation(s). Fix playlist_data.py.",
              file=sys.stderr)
        sys.exit(1)

    all_stats = {pl["name"]: playlist_stats(pl) for pl in PLAYLISTS}
    total_secs = sum(st["seconds"] for st in all_stats.values())
    total_songs = sum(st["songs"] for st in all_stats.values())
    genre_totals = Counter(
        g for pl in PLAYLISTS for _a, _t, _l, g in pl["songs"]
    )

    rows = [
        (pl["name"], n, artist, title, length, genre)
        for pl in PLAYLISTS
        for n, (artist, title, length, genre) in enumerate(pl["songs"], 1)
    ]

    write_master_csv(rows)
    write_per_playlist_csvs()
    write_json()
    write_markdown(all_stats, genre_totals, total_secs, total_songs)
    write_validation_md(errors, warnings, all_stats, genre_totals, total_secs, total_songs)
    write_xlsx(rows, all_stats, total_secs, total_songs)

    print(f"\n{'Playlist':<16} {'Songs':>5} {'Actual':>8} {'Target':>8} {'Drift':>6}")
    for pl in PLAYLISTS:
        st = all_stats[pl["name"]]
        print(f"{pl['name']:<16} {st['songs']:>5} {st['runtime']:>8} "
              f"{st['target']:>8} {st['delta_min']:>+5}m")
    print(f"\nTOTAL: {total_songs} unique songs, {fmt_hm(total_secs)}")
    print("Genre mix: " + ", ".join(
        f"{g} {n * 100 // total_songs}%" for g, n in genre_totals.most_common()))
    print("\nAll hard rules PASSED. Deliverables written to output/")


if __name__ == "__main__":
    main()
